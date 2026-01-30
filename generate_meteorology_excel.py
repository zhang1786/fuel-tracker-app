import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def csv_to_excel(csv_file, excel_file):
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "榆阳区气象局民主生活会谈心谈话记录表"
    
    # 设置标题行样式
    title_font = Font(name='宋体', size=16, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 读取CSV数据
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # 添加表头
    for col_num, cell_value in enumerate(rows[0], 1):
        cell = ws.cell(row=1, column=col_num, value=cell_value)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    # 添加数据行
    for row_num, row_data in enumerate(rows[1:], 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, 12), 50)  # 最小宽度12，最大宽度50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 设置特定列的宽度
    ws.column_dimensions['A'].width = 15  # 谈话时间
    ws.column_dimensions['B'].width = 20  # 谈话地点
    ws.column_dimensions['C'].width = 25  # 谈话人
    ws.column_dimensions['D'].width = 25  # 谈话对象
    ws.column_dimensions['E'].width = 15  # 谈话类型
    ws.column_dimensions['F'].width = 40  # 谈话主要内容
    ws.column_dimensions['G'].width = 30  # 存在问题
    ws.column_dimensions['H'].width = 30  # 整改措施
    ws.column_dimensions['I'].width = 20  # 备注
    
    # 添加表格标题
    ws.insert_rows(1)
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = '榆阳区气象局党员领导干部民主生活会谈心谈话记录表'
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    # 保存Excel文件
    wb.save(excel_file)

# 使用函数
csv_to_excel('temp_meteorology_talk.csv', '榆阳区气象局民主生活会谈心谈话记录表.xlsx')
print("Excel文件已生成：榆阳区气象局民主生活会谈心谈话记录表.xlsx")