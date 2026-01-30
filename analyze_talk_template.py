#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
谈心谈话内容分析模板
用于分析民主生活会、组织生活会前谈心谈话内容
并生成对谈话人和谈话对象的意见建议
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_analysis_template():
    """创建谈心谈话分析模板"""
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "谈心谈话分析报告"
    
    # 设置样式
    title_font = Font(name='宋体', size=16, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建分析模板
    headers = ['项目', '内容']
    ws.append(headers)
    
    # 添加标题
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    ws['A1'].fill = header_fill
    ws['A1'].border = thin_border
    ws['B1'].font = header_font
    ws['B1'].alignment = center_alignment
    ws['B1'].fill = header_fill
    ws['B1'].border = thin_border
    
    # 添加对谈话人的意见建议
    talker_suggestions = [
        "1. 加强与下属的沟通频率，定期开展谈心谈话活动",
        "2. 在谈话中更加注重倾听，充分了解下属的思想动态",
        "3. 提高谈话的针对性，结合实际工作和个人情况",
        "4. 注重谈话效果的跟踪，确保问题得到解决",
        "5. 创新谈话方式，采用多样化的沟通形式",
        "6. 在谈话中加强对政策法规的宣传和解读"
    ]
    
    ws.append(['对谈话人的意见建议', ''])
    for suggestion in talker_suggestions:
        ws.append(['', suggestion])
    
    # 添加对谈话对象的意见建议
    talked_suggestions = [
        "1. 积极主动汇报思想工作情况，增强沟通意识",
        "2. 勇于开展批评与自我批评，直面自身问题",
        "3. 提高政治站位，增强大局意识和责任担当",
        "4. 加强理论学习，不断提升政治素养和业务能力",
        "5. 主动接受监督，虚心听取各方意见建议",
        "6. 注重学以致用，将学习成果转化为工作实效"
    ]
    
    ws.append(['对谈话对象的意见建议', ''])
    for suggestion in talked_suggestions:
        ws.append(['', suggestion])
    
    # 添加列宽设置
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    
    # 应用样式到所有单元格
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.font = header_font
                cell.alignment = center_alignment
                cell.fill = header_fill
            elif cell.column == 1:
                cell.alignment = center_alignment
            else:
                cell.alignment = wrap_alignment
    
    # 保存文件
    wb.save('谈心谈话分析报告模板.xlsx')
    print("谈心谈话分析报告模板已创建：谈心谈话分析报告模板.xlsx")


def create_detailed_analysis_template():
    """创建更详细的分析模板"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "详细谈心谈话分析"
    
    # 设置样式
    title_font = Font(name='宋体', size=16, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建更详细的表格结构
    headers = ['序号', '谈话对象', '谈话时间', '主要问题', '对谈话人的意见建议', '对谈话对象的意见建议']
    ws.append(headers)
    
    # 格式化表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    # 添加示例数据
    sample_data = [
        [1, '张某某', '2024年1月', '工作积极性待提高', '加强激励引导，关注其工作状态', '提高主观能动性，主动承担工作任务'],
        [2, '李某某', '2024年2月', '理论学习不够深入', '定期组织专题学习，加强指导', '制定个人学习计划，提升理论水平'],
        [3, '王某某', '2024年3月', '业务能力有待提升', '安排专业培训，提供实践机会', '主动学习新技术，提升业务技能'],
        [4, '赵某某', '2024年4月', '沟通协调能力不足', '创造更多交流平台，加强锻炼', '积极参与团队协作，提升沟通技巧'],
        [5, '刘某某', '2024年5月', '创新意识不强', '营造创新氛围，鼓励大胆尝试', '解放思想，勇于提出新思路新方法'],
        [6, '陈某某', '2024年6月', '责任担当意识需加强', '明确责任分工，强化责任意识', '敢于担当，主动作为，积极解决问题']
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 15  # 谈话对象
    ws.column_dimensions['C'].width = 15  # 谈话时间
    ws.column_dimensions['D'].width = 25  # 主要问题
    ws.column_dimensions['E'].width = 35  # 对谈话人的意见建议
    ws.column_dimensions['F'].width = 35  # 对谈话对象的意见建议
    
    # 应用样式到数据行
    for row in ws.iter_rows(min_row=2, max_row=len(sample_data)+1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = wrap_alignment
    
    wb.save('详细谈心谈话分析报告.xlsx')
    print("详细谈心谈话分析报告已创建：详细谈心谈话分析报告.xlsx")


if __name__ == "__main__":
    create_analysis_template()
    create_detailed_analysis_template()