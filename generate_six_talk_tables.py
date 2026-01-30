#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成6份不同的党员领导干部民主生活会谈心谈话表
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_talk_table(sheet_number, title, talker_info, talked_info):
    """创建单个谈心谈话表"""
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"谈心谈话表{sheet_number}"
    
    # 设置样式
    title_font = Font(name='宋体', size=16, bold=True)
    header_font = Font(name='宋体', size=12, bold=True)
    normal_font = Font(name='宋体', size=11)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 添加主标题
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f'{title}'
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    # 添加副标题
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = '2025年度'
    subtitle_cell.font = Font(name='宋体', size=14)
    subtitle_cell.alignment = center_alignment
    
    # 添加单位信息
    ws.merge_cells('A3:F3')
    unit_cell = ws['A3']
    unit_cell.value = '单位：榆阳区气象局'
    unit_cell.font = normal_font
    unit_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 添加基本信息行
    ws.append(['谈话人', talker_info['name'], '职务', talker_info['position'], '谈话对象', talked_info['name']])
    ws.append(['', '', '', '', '职务', talked_info['position']])
    ws.append(['', '', '', '', '', ''])
    
    # 格式化基本信息行
    for row in range(4, 7):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col % 2 != 0:  # 奇数列
                cell.alignment = center_alignment
                if row == 4:
                    cell.font = header_font
                    cell.fill = header_fill
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 设置基本信息区域的行高
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 25
    
    # 设置基本信息区域的列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    
    # 添加谈话时间和地点
    ws.merge_cells('A7:C7')
    time_cell = ws['A7']
    time_cell.value = '谈话时间：____年____月____日'
    time_cell.font = normal_font
    time_cell.alignment = center_alignment
    time_cell.border = thin_border
    
    ws.merge_cells('D7:F7')
    loc_cell = ws['D7']
    loc_cell.value = '谈话地点：会议室'
    loc_cell.font = normal_font
    loc_cell.alignment = center_alignment
    loc_cell.border = thin_border
    
    # 添加谈话内容标题
    ws.merge_cells('A8:F8')
    content_title_cell = ws['A8']
    content_title_cell.value = '谈话主要内容'
    content_title_cell.font = header_font
    content_title_cell.alignment = center_alignment
    content_title_cell.fill = header_fill
    content_title_cell.border = thin_border
    
    # 添加谈话内容区域
    ws.merge_cells('A9:F12')
    content_cell = ws['A9']
    content_cell.value = f'（在此区域填写谈话的主要内容）\n\n本次谈话围绕{talked_info["position"]} {talked_info["name"]}同志的思想、工作、作风等方面进行深入交流，了解其思想动态，征求对班子建设的意见建议。'
    content_cell.font = normal_font
    content_cell.alignment = wrap_alignment
    content_cell.border = thin_border
    
    # 设置谈话内容区域高度
    for row in range(9, 13):
        ws.row_dimensions[row].height = 30
    
    # 添加对谈话人的建议标题
    ws.merge_cells('A13:F13')
    suggest_title_cell = ws['A13']
    suggest_title_cell.value = '对谈话人的意见建议（6条）'
    suggest_title_cell.font = header_font
    suggest_title_cell.alignment = center_alignment
    suggest_title_cell.fill = header_fill
    suggest_title_cell.border = thin_border
    
    # 添加对谈话人的建议内容
    talker_suggestions = [
        f"1. 建议{talker_info['name']}同志进一步加强与分管部门干部职工的沟通交流，定期深入一线了解实际情况。",
        f"2. 建议{talker_info['name']}同志在工作中更加注重年轻干部的培养和指导，提供更多学习和锻炼的机会。",
        f"3. 建议{talker_info['name']}同志加强对党风廉政建设的重视，严格落实“一岗双责”，筑牢思想防线。",
        f"4. 建议{talker_info['name']}同志在决策过程中更加注重民主集中制原则，广泛征求各方意见，确保决策科学合理。",
        f"5. 建议{talker_info['name']}同志进一步改进工作作风，提高工作效率，特别是在解决实际问题方面加大力度。",
        f"6. 建议{talker_info['name']}同志加强理论学习，不断提升政治素养和业务能力，更好地适应新时代工作要求。"
    ]
    
    for i, suggestion in enumerate(talker_suggestions, start=14):
        ws.merge_cells(f'A{i}:F{i}')
        cell = ws['A' + str(i)]
        cell.value = suggestion
        cell.font = normal_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        ws.row_dimensions[i].height = 25
    
    # 添加对谈话对象的建议标题
    ws.merge_cells('A20:F20')
    talked_title_cell = ws['A20']
    talked_title_cell.value = '对谈话对象的意见建议（6条）'
    talked_title_cell.font = header_font
    talked_title_cell.alignment = center_alignment
    talked_title_cell.fill = header_fill
    talked_title_cell.border = thin_border
    
    # 添加对谈话对象的建议内容
    talked_suggestions = [
        f"1. 建议{talked_info['name']}同志进一步加强政治理论学习，提升政治站位，增强“四个意识”，坚定“四个自信”，做到“两个维护”。",
        f"2. 建议{talked_info['name']}同志在工作中主动担当作为，面对困难和挑战时不退缩，积极寻求解决办法。",
        f"3. 建议{talked_info['name']}同志进一步增强服务意识，提高服务质量，更好地为人民群众和地方经济社会发展服务。",
        f"4. 建议{talked_info['name']}同志加强业务学习，不断提升专业能力和技术水平，适应气象现代化发展需要。",
        f"5. 建议{talked_info['name']}同志严格遵守廉洁自律各项规定，筑牢拒腐防变的思想防线，始终保持清正廉洁。",
        f"6. 建议{talked_info['name']}同志进一步改进工作作风，提高工作效率，确保各项工作任务按时保质完成。"
    ]
    
    for i, suggestion in enumerate(talked_suggestions, start=21):
        ws.merge_cells(f'A{i}:F{i}')
        cell = ws['A' + str(i)]
        cell.value = suggestion
        cell.font = normal_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        ws.row_dimensions[i].height = 25
    
    # 添加问题及建议区域
    ws.merge_cells('A27:F27')
    issue_title_cell = ws['A27']
    issue_title_cell.value = '存在问题及改进措施'
    issue_title_cell.font = header_font
    issue_title_cell.alignment = center_alignment
    issue_title_cell.fill = header_fill
    issue_title_cell.border = thin_border
    
    ws.merge_cells('A28:F31')
    issue_cell = ws['A28']
    issue_cell.value = f'{talker_info["position"]}{talker_info["name"]}同志存在问题及改进措施：\n\n{talked_info["position"]}{talked_info["name"]}同志存在问题及改进措施：'
    issue_cell.font = normal_font
    issue_cell.alignment = wrap_alignment
    issue_cell.border = thin_border
    
    for row in range(28, 32):
        ws.row_dimensions[row].height = 25
    
    # 添加签名区域
    ws.merge_cells('A32:C32')
    signer1_cell = ws['A32']
    signer1_cell.value = '谈话人签字：           年    月    日'
    signer1_cell.font = normal_font
    signer1_cell.alignment = Alignment(horizontal='left', vertical='center')
    signer1_cell.border = thin_border
    
    ws.merge_cells('D32:F32')
    signer2_cell = ws['D32']
    signer2_cell.value = '谈话对象签字：         年    月    日'
    signer2_cell.font = normal_font
    signer2_cell.alignment = Alignment(horizontal='left', vertical='center')
    signer2_cell.border = thin_border
    
    ws.row_dimensions[32].height = 25
    
    # 保存文件
    filename = f'谈心谈话表{sheet_number}-{talked_info["name"]}.xlsx'
    wb.save(filename)
    print(f"谈心谈话表{sheet_number}已生成：{filename}")
    return filename


def generate_all_tables():
    """生成6份不同的谈心谈话表"""
    
    # 定义不同的谈话场景
    scenarios = [
        {
            "title": "党组书记与党组成员副局长谈心谈话表",
            "talker_info": {"name": "张伟", "position": "党组书记、局长"},
            "talked_info": {"name": "李华", "position": "党组成员、副局长"}
        },
        {
            "title": "党组书记与纪检组长谈心谈话表",
            "talker_info": {"name": "张伟", "position": "党组书记、局长"},
            "talked_info": {"name": "赵敏", "position": "党组成员、纪检组长"}
        },
        {
            "title": "党组成员与科室主任谈心谈话表",
            "talker_info": {"name": "李华", "position": "党组成员、副局长"},
            "talked_info": {"name": "刘晓东", "position": "监测预警科科长"}
        },
        {
            "title": "党组成员与技术骨干谈心谈话表",
            "talker_info": {"name": "王建国", "position": "党组成员、副局长"},
            "talked_info": {"name": "陈丽", "position": "公共服务科科长"}
        },
        {
            "title": "纪检组长与办公室主任谈心谈话表",
            "talker_info": {"name": "赵敏", "position": "党组成员、纪检组长"},
            "talked_info": {"name": "马涛", "position": "办公室主任"}
        },
        {
            "title": "副局长与年轻干部谈心谈话表",
            "talker_info": {"name": "李华", "position": "党组成员、副局长"},
            "talked_info": {"name": "王小军", "position": "预报员"}
        }
    ]
    
    filenames = []
    for i, scenario in enumerate(scenarios, 1):
        filename = create_talk_table(i, scenario["title"], scenario["talker_info"], scenario["talked_info"])
        filenames.append(filename)
    
    print("\n所有谈心谈话表已生成完毕！")
    return filenames


if __name__ == "__main__":
    generated_files = generate_all_tables()