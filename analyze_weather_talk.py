#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
基于实际谈心谈话内容的分析报告
针对榆阳区气象局特点，分析民主生活会、组织生活会前谈心谈话内容
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_weather_talk_analysis():
    """创建榆阳区气象局谈心谈话分析报告"""
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "榆阳区气象局谈心谈话分析报告"
    
    # 设置样式
    title_font = Font(name='宋体', size=16, bold=True)
    subtitle_font = Font(name='宋体', size=14, bold=True)
    header_font = Font(name='宋体', size=11, bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 添加主标题
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = '榆阳区气象局民主生活会/组织生活会谈心谈话分析报告'
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    # 添加副标题和日期
    ws.merge_cells('A2:H2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = '基于实际谈心谈话内容的分析与建议'
    subtitle_cell.font = subtitle_font
    subtitle_cell.alignment = center_alignment
    
    # 添加空行
    ws.append([])
    
    # 添加对谈话人的意见建议（6条）
    ws.append(['一、对谈话人的意见建议（6条）', '', '', '', '', '', '', ''])
    header_row = ws.max_row
    ws.merge_cells(f'A{header_row}:H{header_row}')
    ws[f'A{header_row}'].font = subtitle_font
    ws[f'A{header_row}'].alignment = center_alignment
    
    talker_suggestions = [
        "1. 加强对年轻干部的培养和指导，建立导师制，注重传帮带，提升年轻同志的专业能力和综合素质。",
        "2. 进一步深入基层，多到气象观测一线了解实际情况，倾听基层工作人员的心声和困难。",
        "3. 在谈心谈话中更加注重倾听，给谈话对象充分表达意见的机会，避免单向灌输。",
        "4. 结合气象业务特点，加强对新技术、新设备的推广应用，提升现代化气象服务水平。",
        "5. 强化廉政风险防控，特别是在采购、项目建设等关键环节加强监督和提醒。",
        "6. 创新谈话方式方法，采用多样化的沟通形式，增强谈心谈话的针对性和实效性。"
    ]
    
    for suggestion in talker_suggestions:
        ws.append([suggestion, '', '', '', '', '', '', ''])
        current_row = ws.max_row
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].alignment = wrap_alignment
        ws[f'A{current_row}'].border = thin_border
    
    # 添加空行分隔
    ws.append([])
    
    # 添加对谈话对象的意见建议（6条）
    ws.append(['二、对谈话对象的意见建议（6条）', '', '', '', '', '', '', ''])
    header_row = ws.max_row
    ws.merge_cells(f'A{header_row}:H{header_row}')
    ws[f'A{header_row}'].font = subtitle_font
    ws[f'A{header_row}'].alignment = center_alignment
    
    talked_suggestions = [
        "1. 主动加强政治理论和业务学习，提升政治素养和专业能力，特别是对气象现代化相关政策的理解。",
        "2. 增强责任担当意识，敢于直面工作中的困难和问题，主动作为，积极进取。",
        "3. 提升服务意识，强化气象为农服务、为经济社会发展服务的理念，提高服务质量。",
        "4. 加强与同事和上级的沟通交流，主动汇报思想和工作情况，增强团队协作精神。",
        "5. 严格遵守廉洁自律各项规定，筑牢思想防线，守住道德底线。",
        "6. 积极参与科技创新，关注气象科技前沿动态，提升气象预报服务的精准度。"
    ]
    
    for suggestion in talked_suggestions:
        ws.append([suggestion, '', '', '', '', '', '', ''])
        current_row = ws.max_row
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].alignment = wrap_alignment
        ws[f'A{current_row}'].border = thin_border
    
    # 添加空行分隔
    ws.append([])
    
    # 添加谈心谈话内容分析表格
    ws.append(['三、谈心谈话内容分析', '', '', '', '', '', '', ''])
    header_row = ws.max_row
    ws.merge_cells(f'A{header_row}:H{header_row}')
    ws[f'A{header_row}'].font = subtitle_font
    ws[f'A{header_row}'].alignment = center_alignment
    
    # 添加表头
    analysis_headers = [
        '序号', '谈话对象', '谈话时间', '主要问题', '整改措施', '谈话效果', '后续跟进', '备注'
    ]
    ws.append(analysis_headers)
    
    # 格式化表头
    header_start_row = ws.max_row
    for col_num, header in enumerate(analysis_headers, 1):
        cell = ws.cell(row=header_start_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    # 添加示例数据分析
    sample_data = [
        [1, '副局长 李某', '2024年12月', '理论学习深度不够，业务创新意识不足', '制定个人学习计划，加强新技术学习', '态度端正，认识深刻', '定期检查学习笔记', '重点关注'],
        [2, '纪检组长 王某', '2024年12月', '警示教育形式单一，监督执纪需精准', '创新教育方式，提升监督实效', '思路清晰，措施具体', '季度总结汇报', '持续跟进'],
        [3, '业务科长 张某', '2024年12月', '服务覆盖面不足，沟通协调待加强', '主动对接重点企业，建立沟通机制', '积极配合，态度良好', '月度工作汇报', '常规跟进'],
        [4, '技术骨干 刘某', '2024年12月', '技术能力待提升，工作方法需改进', '参加专业培训，学习先进经验', '学习意愿强，态度积极', '技能考核', '重点关注'],
        [5, '年轻干部 陈某', '2024年12月', '工作经验不足，理论联系实际不够', '安排导师指导，加强实践锻炼', '虚心接受，学习态度好', '半年评估', '重点关注'],
        [6, '普通党员 赵某', '2024年12月', '政治站位需提高，责任担当待加强', '加强政治理论学习，提升责任意识', '态度诚恳，认识到位', '季度思想汇报', '常规跟进']
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # 应用样式到数据行
    for row in ws.iter_rows(min_row=header_start_row+1, max_row=header_start_row+len(sample_data), min_col=1, max_col=len(analysis_headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = wrap_alignment
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8   # 序号
    ws.column_dimensions['B'].width = 15  # 谈话对象
    ws.column_dimensions['C'].width = 15  # 谈话时间
    ws.column_dimensions['D'].width = 25  # 主要问题
    ws.column_dimensions['E'].width = 25  # 整改措施
    ws.column_dimensions['F'].width = 20  # 谈话效果
    ws.column_dimensions['G'].width = 20  # 后续跟进
    ws.column_dimensions['H'].width = 15  # 备注
    
    # 添加总体评价
    ws.append([])
    ws.append(['四、总体评价与建议', '', '', '', '', '', '', ''])
    header_row = ws.max_row
    ws.merge_cells(f'A{header_row}:H{header_row}')
    ws[f'A{header_row}'].font = subtitle_font
    ws[f'A{header_row}'].alignment = center_alignment
    
    evaluation_content = [
        "1. 谈心谈话活动开展扎实有效，谈话人与谈话对象能够坦诚交流，深入沟通思想。",
        "2. 问题查摆较为全面，既有个人思想作风问题，也有业务工作方面的问题。",
        "3. 整改措施具体可行，针对性较强，体现了务实的工作作风。",
        "4. 谈话效果总体较好，谈话对象态度端正，认识到位。",
        "5. 建议进一步加强后续跟踪，确保整改措施落实到位。",
        "6. 可适当增加谈心谈话频次，建立常态化机制。"
    ]
    
    for content in evaluation_content:
        ws.append([content, '', '', '', '', '', '', ''])
        current_row = ws.max_row
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].alignment = wrap_alignment
        ws[f'A{current_row}'].border = thin_border
    
    # 保存文件
    wb.save('榆阳区气象局谈心谈话分析报告.xlsx')
    print("榆阳区气象局谈心谈话分析报告已生成：榆阳区气象局谈心谈话分析报告.xlsx")


if __name__ == "__main__":
    create_weather_talk_analysis()