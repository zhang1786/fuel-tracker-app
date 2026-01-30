import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def csv_to_excel(csv_file, excel_file):
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "谈心谈话记录"
    
    # 读取CSV数据
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # 设置标题行样式
    for col_num, cell_value in enumerate(rows[0], 1):
        cell = ws.cell(row=1, column=col_num, value=cell_value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加数据行
    for row_num, row_data in enumerate(rows[1:], 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存Excel文件
    wb.save(excel_file)

# 使用函数
csv_to_excel('temp_talk_content.csv', '谈心谈话记录.xlsx')
print("Excel文件已生成：谈心谈话记录.xlsx")